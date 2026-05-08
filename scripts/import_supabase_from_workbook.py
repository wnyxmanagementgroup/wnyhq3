#!/usr/bin/env python3
"""
นำเข้าข้อมูลจากไฟล์ Excel ของระบบไปราชการไปยัง Supabase

รองรับชีต:
- Users
- Requests
- Attendees
- Memos
- Trash

ตัวอย่างการใช้งาน:
python3 scripts/import_supabase_from_workbook.py \
  --xlsx "/path/to/workbook.xlsx" \
  --url "https://YOUR_PROJECT.supabase.co" \
  --service-role-key "YOUR_SERVICE_ROLE_KEY"
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import workbook data into Supabase")
    parser.add_argument("--xlsx", required=True, help="พาธไฟล์ Excel ที่จะนำเข้า")
    parser.add_argument("--url", default=os.getenv("SUPABASE_URL"), help="Supabase project URL")
    parser.add_argument(
        "--service-role-key",
        default=os.getenv("SUPABASE_SERVICE_ROLE_KEY"),
        help="Supabase service role key",
    )
    parser.add_argument("--schema", default="public", help="ชื่อ schema ของ PostgREST")
    parser.add_argument("--chunk-size", type=int, default=200, help="จำนวนแถวต่อการส่งหนึ่งครั้ง")
    parser.add_argument("--dry-run", action="store_true", help="แสดงจำนวนข้อมูลโดยไม่อัปโหลดจริง")
    return parser.parse_args()


def to_iso_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def to_iso_timestamp(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            pass
    return text


def to_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_numeric(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_bool(value: Any) -> Optional[bool]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def to_json_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (dict, list)):
        return value
    text = str(value).strip()
    if not text:
        return None
    if text in {"[]", "{}"}:
        return json.loads(text)
    if text.startswith("[") or text.startswith("{"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def normalize_headers(row: Iterable[Any]) -> List[str]:
    headers: List[str] = []
    for idx, cell in enumerate(row, start=1):
        text = str(cell).strip() if cell is not None else ""
        headers.append(text or f"__blank_{idx}")
    return headers


def sheet_records(workbook_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = normalize_headers(rows[0])
    records: List[Dict[str, Any]] = []
    for row_index, values in enumerate(rows[1:], start=2):
        if values is None:
            continue
        record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if all(v is None or str(v).strip() == "" for v in record.values()):
            continue
        record["__row_index"] = row_index
        records.append(record)
    return records


def map_users(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in rows:
        username = to_text(row.get("Username"))
        if not username:
            continue
        result.append(
            {
                "username": username,
                "login_name": to_text(row.get("LoginName")),
                "full_name": to_text(row.get("FullName")),
                "position": to_text(row.get("Position")),
                "department": to_text(row.get("Department")),
                "role": to_text(row.get("Role")),
                "email": to_text(row.get("Email")),
                "special_position": to_text(row.get("SpecialPosition")),
                "token": to_text(row.get("Token")),
                "legacy_password": to_text(row.get("Password")),
            }
        )
    return result


def map_requests(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in rows:
        request_id = to_text(row.get("RequestId"))
        if not request_id:
            continue
        result.append(
            {
                "request_id": request_id,
                "created_by": to_text(row.get("CreatedBy")),
                "ref_number": to_text(row.get("RefNumber")),
                "doc_date": to_iso_date(row.get("DocDate")),
                "requester_name": to_text(row.get("RequesterName")),
                "requester_position": to_text(row.get("RequesterPosition")),
                "location": to_text(row.get("Location")),
                "purpose": to_text(row.get("Purpose")),
                "start_date": to_iso_date(row.get("StartDate")),
                "end_date": to_iso_date(row.get("EndDate")),
                "duration": to_text(row.get("Duration")),
                "expense_option": to_text(row.get("ExpenseOption")),
                "expense_items": to_json_value(row.get("ExpenseItems")),
                "total_expense": to_numeric(row.get("TotalExpense")),
                "vehicle_option": to_text(row.get("VehicleOption")),
                "license_plate": to_text(row.get("LicensePlate")),
                "department": to_text(row.get("Department")),
                "head_name": to_text(row.get("HeadName")),
                "pdf_url": to_text(row.get("PdfUrl")),
                "created_at_source": to_iso_timestamp(row.get("Timestamp")),
                "status": to_text(row.get("Status")),
                "command_pdf_url": to_text(row.get("CommandPdfUrl")),
                "command_status": to_text(row.get("CommandStatus")),
                "command_pdf_url_solo": to_text(row.get("CommandPdfUrlSolo")),
                "command_pdf_url_group_small": to_text(row.get("CommandPdfUrlGroupSmall")),
                "command_pdf_url_group_large": to_text(row.get("CommandPdfUrlGroupLarge")),
                "dispatch_book_pdf_url": to_text(row.get("DispatchBookPdfUrl")),
                "item_1": to_text(row.get("Item1")),
                "qty_1": to_numeric(row.get("Qty1")),
                "item_2": to_text(row.get("Item2")),
                "qty_2": to_numeric(row.get("Qty2")),
                "item_3": to_text(row.get("Item3")),
                "qty_3": to_numeric(row.get("Qty3")),
                "item_4": to_text(row.get("Item4")),
                "qty_4": to_numeric(row.get("Qty4")),
                "item_5": to_text(row.get("Item5")),
                "qty_5": to_numeric(row.get("Qty5")),
                "item_6": to_text(row.get("Item6")),
                "qty_6": to_numeric(row.get("Qty6")),
                "item_7": to_text(row.get("Item7")),
                "qty_7": to_numeric(row.get("Qty7")),
                "command_doc_url_solo": to_text(row.get("CommandDocUrlSolo")),
                "command_doc_url_group_large": to_text(row.get("CommandDocUrlGroupLarge")),
                "doc_url": to_text(row.get("DocUrl")),
                "command_doc_url_group_small": to_text(row.get("CommandDocUrlGroupSmall")),
                "province": to_text(row.get("Province")),
                "stay_at": to_text(row.get("StayAt")),
                "dispatch_vehicle_type": to_text(row.get("DispatchVehicleType")),
                "dispatch_vehicle_id": to_text(row.get("DispatchVehicleId")),
                "completed_memo_url": to_text(row.get("CompletedMemoUrl")),
                "completed_command_url": to_text(row.get("CompletedCommandUrl")),
                "memo_status": to_text(row.get("MemoStatus")),
                "dispatch_book_url": to_text(row.get("DispatchBookUrl")),
                "admin_memo_url": to_text(row.get("AdminMemoUrl")),
                "doc_status": to_text(row.get("DocStatus")),
                "was_rejected": to_bool(row.get("WasRejected")),
                "rejection_reason": to_text(row.get("RejectionReason")),
                "command_template_type": to_text(row.get("CommandTemplateType")),
            }
        )
    return result


def map_attendees(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in rows:
        request_id = to_text(row.get("RequestId"))
        full_name = to_text(row.get("FullName"))
        if not request_id or not full_name:
            continue
        row_index = row.get("__row_index")
        result.append(
            {
                "source_row_key": f"attendees:{row_index}",
                "request_id": request_id,
                "full_name": full_name,
                "position": to_text(row.get("Position")),
                "source_date_text": to_text(row.get("Date")),
                "attended_at": to_iso_timestamp(row.get("Date")),
            }
        )
    return result


def map_memos(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in rows:
        memo_id = to_text(row.get("MemoID"))
        if not memo_id:
            continue
        result.append(
            {
                "memo_id": memo_id,
                "submitted_by": to_text(row.get("SubmittedBy")),
                "ref_number": to_text(row.get("RefNumber")),
                "status": to_text(row.get("Status")),
                "created_at_source": to_iso_timestamp(row.get("Timestamp")),
                "file_id": to_text(row.get("FileID")),
                "file_url": to_text(row.get("FileURL")),
                "completed_memo_url": to_text(row.get("CompletedMemoUrl")),
                "completed_command_url": to_text(row.get("CompletedCommandUrl")),
                "dispatch_book_url": to_text(row.get("DispatchBookUrl")),
            }
        )
    return result


def map_trash(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in rows:
        request_id = to_text(row.get("RequestId"))
        row_index = row.get("__row_index")
        if not request_id and not row_index:
            continue
        result.append(
            {
                "source_row_key": f"trash:{row_index}",
                "request_id": request_id,
                "created_by": to_text(row.get("CreatedBy")),
                "ref_number": to_text(row.get("RefNumber")),
                "doc_date": to_iso_date(row.get("DocDate")),
                "requester_name": to_text(row.get("RequesterName")),
                "requester_position": to_text(row.get("RequesterPosition")),
                "location": to_text(row.get("Location")),
                "purpose": to_text(row.get("Purpose")),
                "start_date": to_iso_date(row.get("StartDate")),
                "end_date": to_iso_date(row.get("EndDate")),
                "duration": to_text(row.get("Duration")),
                "expense_option": to_text(row.get("ExpenseOption")),
                "expense_items": to_json_value(row.get("ExpenseItems")),
                "total_expense": to_numeric(row.get("TotalExpense")),
                "vehicle_option": to_text(row.get("VehicleOption")),
                "license_plate": to_text(row.get("LicensePlate")),
                "department": to_text(row.get("Department")),
                "head_name": to_text(row.get("HeadName")),
                "pdf_url": to_text(row.get("PdfUrl")),
                "created_at_source": to_iso_timestamp(row.get("Timestamp")),
                "status": to_text(row.get("Status")),
                "command_pdf_url": to_text(row.get("CommandPdfUrl")),
                "command_status": to_text(row.get("CommandStatus")),
                "command_pdf_url_solo": to_text(row.get("CommandPdfUrlSolo")),
                "command_pdf_url_group_small": to_text(row.get("CommandPdfUrlGroupSmall")),
                "command_pdf_url_group_large": to_text(row.get("CommandPdfUrlGroupLarge")),
                "dispatch_book_pdf_url": to_text(row.get("DispatchBookPdfUrl")),
                "command_doc_url_solo": to_text(row.get("CommandDocUrlSolo")),
                "command_doc_url_group_large": to_text(row.get("CommandDocUrlGroupLarge")),
                "doc_url": to_text(row.get("DocUrl")),
                "command_doc_url_group_small": to_text(row.get("CommandDocUrlGroupSmall")),
                "province": to_text(row.get("Province")),
                "stay_at": to_text(row.get("StayAt")),
                "dispatch_vehicle_type": to_text(row.get("DispatchVehicleType")),
                "dispatch_vehicle_id": to_text(row.get("DispatchVehicleId")),
                "completed_memo_url": to_text(row.get("CompletedMemoUrl")),
                "completed_command_url": to_text(row.get("CompletedCommandUrl")),
                "deleted_at": to_iso_timestamp(row.get("DeletedAt")),
                "deleted_by": to_text(row.get("DeletedBy")),
            }
        )
    return result


def build_request_counters(requests_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    counter_by_year: Dict[int, int] = defaultdict(int)
    pattern = re.compile(r"^บค(\d+)/(\d{4})$")
    for row in requests_rows:
        request_id = to_text(row.get("request_id"))
        if not request_id:
            continue
        match = pattern.match(request_id)
        if not match:
            continue
        seq = int(match.group(1))
        year_be = int(match.group(2))
        counter_by_year[year_be] = max(counter_by_year[year_be], seq)
    return [
        {"year_be": year_be, "current_count": count}
        for year_be, count in sorted(counter_by_year.items())
    ]


def postgrest_upsert(
    *,
    base_url: str,
    schema: str,
    table: str,
    rows: List[Dict[str, Any]],
    chunk_size: int,
    api_key: str,
    on_conflict: str,
) -> None:
    if not rows:
        print(f"- ข้าม {table}: ไม่มีข้อมูล")
        return

    endpoint = (
        f"{base_url.rstrip('/')}/rest/v1/{table}"
        f"?on_conflict={urllib.parse.quote(on_conflict)}"
    )
    headers = {
        "Content-Type": "application/json",
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Prefer": "resolution=merge-duplicates,return=minimal",
        "Accept-Profile": schema,
        "Content-Profile": schema,
    }

    total = len(rows)
    for start in range(0, total, chunk_size):
        chunk = rows[start : start + chunk_size]
        payload = json.dumps(chunk, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(endpoint, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(request) as response:
                if response.status not in (200, 201, 204):
                    raise RuntimeError(f"HTTP {response.status}")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"อัปโหลด {table} ไม่สำเร็จ: HTTP {exc.code} {body}") from exc
        print(f"- อัปโหลด {table}: {min(start + len(chunk), total)}/{total}")


def main() -> int:
    args = parse_args()
    if not args.url or not args.service_role_key:
        print("กรุณาระบุ --url และ --service-role-key หรือกำหนดผ่าน environment variables", file=sys.stderr)
        return 1

    workbook_path = args.xlsx
    if not os.path.exists(workbook_path):
        print(f"ไม่พบไฟล์: {workbook_path}", file=sys.stderr)
        return 1

    users = map_users(sheet_records(workbook_path, "Users"))
    requests = map_requests(sheet_records(workbook_path, "Requests"))
    attendees = map_attendees(sheet_records(workbook_path, "Attendees"))
    memos = map_memos(sheet_records(workbook_path, "Memos"))
    trash_requests = map_trash(sheet_records(workbook_path, "Trash"))
    request_counters = build_request_counters(requests)

    summary = {
        "app_users": len(users),
        "requests": len(requests),
        "attendees": len(attendees),
        "memos": len(memos),
        "trash_requests": len(trash_requests),
        "request_counters": len(request_counters),
    }

    print("สรุปจำนวนข้อมูลที่จะนำเข้า")
    for table, count in summary.items():
        print(f"- {table}: {count}")

    if args.dry_run:
        print("โหมด dry-run: ยังไม่ได้อัปโหลดข้อมูลจริง")
        return 0

    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="app_users",
        rows=users,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="username",
    )
    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="requests",
        rows=requests,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="request_id",
    )
    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="attendees",
        rows=attendees,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="source_row_key",
    )
    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="memos",
        rows=memos,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="memo_id",
    )
    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="trash_requests",
        rows=trash_requests,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="source_row_key",
    )
    postgrest_upsert(
        base_url=args.url,
        schema=args.schema,
        table="request_counters",
        rows=request_counters,
        chunk_size=args.chunk_size,
        api_key=args.service_role_key,
        on_conflict="year_be",
    )

    print("นำเข้าข้อมูลเสร็จเรียบร้อย")
    print("หมายเหตุ: approval_links, app_settings และ system_config ต้อง seed เพิ่มต่างหาก")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
